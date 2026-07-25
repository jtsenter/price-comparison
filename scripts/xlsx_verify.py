"""Independent verification of the workbook produced by docs/xlsx.js.

Reads the file with Python's OWN zipfile + XML parser, so a bug in our
hand-rolled ZIP writer cannot be masked by a matching bug in our reader.
This is the check that actually stands in for "will Excel open it?".

Run after scripts/xlsx_selfcheck.js (which writes the file):
    node scripts/xlsx_selfcheck.js && python scripts/xlsx_verify.py
"""
import os
import sys
import tempfile
import zipfile
import xml.etree.ElementTree as ET

PATH = os.path.join(tempfile.gettempdir(), "pw_xlsx_selfcheck.xlsx")
NS = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}

REQUIRED = [
    "[Content_Types].xml",
    "_rels/.rels",
    "xl/workbook.xml",
    "xl/_rels/workbook.xml.rels",
    "xl/styles.xml",
    "xl/worksheets/sheet1.xml",
]


def main() -> int:
    if not os.path.exists(PATH):
        print(f"FAIL: {PATH} not found - run `node scripts/xlsx_selfcheck.js` first")
        return 1

    checks = 0
    if not zipfile.is_zipfile(PATH):
        print("FAIL: not a valid ZIP archive")
        return 1
    checks += 1

    with zipfile.ZipFile(PATH) as z:
        bad = z.testzip()
        if bad is not None:
            print(f"FAIL: CRC mismatch in {bad}")
            return 1
        checks += 1

        names = z.namelist()
        for part in REQUIRED:
            if part not in names:
                print(f"FAIL: missing part {part}")
                return 1
            checks += 1

        # Every part must be well-formed XML - this is what Excel chokes on.
        for part in REQUIRED:
            try:
                ET.fromstring(z.read(part))
            except ET.ParseError as e:
                print(f"FAIL: {part} is not well-formed XML: {e}")
                return 1
            checks += 1

        sheet = ET.fromstring(z.read("xl/worksheets/sheet1.xml"))
        rows = sheet.findall(".//m:sheetData/m:row", NS)
        if len(rows) != 4:  # 1 header + 3 data
            print(f"FAIL: expected 4 rows, got {len(rows)}")
            return 1
        checks += 1

        def cell_text(c):
            t = c.find("m:is/m:t", NS)
            if t is not None:
                return t.text
            v = c.find("m:v", NS)
            return v.text if v is not None else None

        header = [cell_text(c) for c in rows[0].findall("m:c", NS)]
        expected = ["Date", "Time", "Product", "Previous price", "New price", "Store"]
        if header != expected:
            print(f"FAIL: header mismatch\n  got      {header}\n  expected {expected}")
            return 1
        checks += 1

        r2 = rows[1].findall("m:c", NS)
        # Prices must be numeric cells (no t attr) so Excel can sum/pivot them.
        if r2[3].get("t") is not None:
            print("FAIL: price cell is text, not a number")
            return 1
        if abs(float(cell_text(r2[3])) - 3.85) > 1e-9:
            print(f"FAIL: price value wrong: {cell_text(r2[3])}")
            return 1
        checks += 2

        # Date cell numeric + date-styled, and round-trips to the right day.
        if r2[0].get("t") is not None:
            print("FAIL: date cell is text, not a number")
            return 1
        serial = float(cell_text(r2[0]))
        from datetime import datetime, timedelta
        got = datetime(1899, 12, 30) + timedelta(days=serial)
        if (got.year, got.month, got.day) != (2026, 7, 22):
            print(f"FAIL: date serial {serial} -> {got}, expected 2026-07-22")
            return 1
        checks += 2

        # The XML-hostile product name must survive intact after parsing.
        r3 = rows[2].findall("m:c", NS)
        name = cell_text(r3[2])
        if name != 'Ben & Jerry\'s <"Half Baked"> 458ml':
            print(f"FAIL: escaped text did not round-trip: {name!r}")
            return 1
        checks += 1

        wb = ET.fromstring(z.read("xl/workbook.xml"))
        sheets = wb.findall(".//m:sheets/m:sheet", NS)
        if len(sheets) != 1 or sheets[0].get("name") != "Price changes":
            print("FAIL: expected exactly one sheet named 'Price changes'")
            return 1
        checks += 1

    # openpyxl is the real-world reader most people use - check it too if present.
    try:
        import openpyxl  # type: ignore
        wb = openpyxl.load_workbook(PATH)
        ws = wb.active
        assert ws.max_row == 4 and ws.max_column == 6
        assert ws.cell(2, 4).value == 3.85
        print("  openpyxl: opened cleanly, values typed correctly")
        checks += 1
    except ImportError:
        print("  openpyxl not installed - skipped (zipfile+XML checks still ran)")

    print(f"xlsx_verify: all {checks} checks passed - {PATH} is a valid workbook")
    return 0


if __name__ == "__main__":
    sys.exit(main())
